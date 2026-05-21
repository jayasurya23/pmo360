"""
Parse a Castillo project schedule from one of two source formats:

  - **xlsx**: a "duration" workbook with a `Project Info` sheet
    (Project Name, Project Start Date, Proposal Version, ...) and a `Tasks`
    sheet whose rows have: ID, Name, Duration, Price, Is Milestone, Indent
    Level, Parent ID, Predecessor ID, Lag. Dates are computed from the start
    date by walking finish-to-start relationships (5-day workweek).

  - **pdf**: the rendered proposal "Project Schedule & Schedule of Value" page
    with columns Task / Days / Start / Finish / Price. Tables are extracted
    with pdfplumber.

Both return a `ParsedSchedule` Pydantic model that the UI saves as
Schedule + ScheduleItem rows.

The discipline / phase / task hierarchy is recovered from indent levels
(xlsx) or row styling cues / known discipline labels (pdf).
"""
from __future__ import annotations

import io
import re
from datetime import date, timedelta
from typing import Optional

from pydantic import BaseModel
from openpyxl import load_workbook
import pdfplumber


# Names that count as top-level discipline rows in both formats
DISCIPLINE_NAMES = {
    "Project Initiation", "Project Closeout",
    "Civil Engineering", "Electrical Engineering", "Structural Engineering",
}
# Names that count as phase rows. Matches the WHOLE task name, so a sub-task
# like "30% - Planset/ Basis of Design" is NOT confused with the "30% Design"
# header above it.
PHASE_PATTERN = re.compile(
    r"^(?:\d+%\s+Design|IFC\s+Design|Studies\s+Updates|IFR(?:\s+Design)?)\s*$",
    re.IGNORECASE,
)


# ============================================================
# Pydantic schema
# ============================================================
class ParsedScheduleItem(BaseModel):
    order_index: int
    indent_level: int                       # 0 = discipline, 1 = phase, 2 = task
    discipline: str = ""                    # filled for items under a discipline
    phase: str = ""                         # filled for items under a phase
    task: str
    duration_days: Optional[int] = None
    start_date: Optional[date] = None
    finish_date: Optional[date] = None
    price: Optional[int] = None             # whole dollars
    is_milestone: bool = False


class ParsedSchedule(BaseModel):
    version: str = "V1"
    source_format: str                      # "pdf" or "xlsx"
    source_filename: str = ""
    project_name: str = ""
    project_start_date: Optional[date] = None
    total_duration_days: Optional[int] = None
    total_price: Optional[int] = None
    items: list[ParsedScheduleItem]


# ============================================================
# Public entry point
# ============================================================
def parse_schedule_file(filename: str, content: bytes) -> ParsedSchedule:
    """Route to the appropriate parser based on filename extension."""
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return parse_excel_schedule(content, filename=filename)
    if name.endswith(".pdf"):
        return parse_pdf_schedule(content, filename=filename)
    raise ValueError(
        f"Unsupported schedule format: {filename!r}. Use .xlsx or .pdf."
    )


# ============================================================
# Excel parser
# ============================================================
def parse_excel_schedule(content: bytes, filename: str = "") -> ParsedSchedule:
    wb = load_workbook(io.BytesIO(content), data_only=True)

    # --- Project Info sheet ---
    info = {}
    if "Project Info" in wb.sheetnames:
        ws_info = wb["Project Info"]
        for row in ws_info.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            key = str(row[0]).strip()
            val = row[1]
            info[key] = val

    project_name = str(info.get("Project Name") or "").strip()
    version = str(info.get("Proposal Version") or "V1").strip() or "V1"
    project_start = _coerce_date(info.get("Project Start Date"))

    # --- Tasks sheet ---
    if "Tasks" not in wb.sheetnames:
        raise ValueError("Excel workbook missing required 'Tasks' sheet.")
    ws = wb["Tasks"]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if len(rows) < 2:
        return ParsedSchedule(
            version=version, source_format="xlsx", source_filename=filename,
            project_name=project_name, project_start_date=project_start, items=[],
        )

    header = [str(c or "").strip() for c in rows[0]]
    idx = {name: i for i, name in enumerate(header)}
    required = {"ID", "Name", "Duration", "Indent Level"}
    missing = required - set(idx)
    if missing:
        raise ValueError(f"Excel 'Tasks' sheet missing columns: {sorted(missing)}")

    # Read raw task records
    raw_tasks: list[dict] = []
    for r in rows[1:]:
        if r is None or r[idx["ID"]] is None:
            continue
        raw_tasks.append({
            "id": int(r[idx["ID"]]),
            "name": str(r[idx["Name"]] or "").strip(),
            "duration": _coerce_int(r[idx["Duration"]]) or 0,
            "price": _coerce_int(r[idx["Price"]]) if "Price" in idx else None,
            "is_milestone": bool(r[idx["Is Milestone"]]) if "Is Milestone" in idx else False,
            "indent": _coerce_int(r[idx["Indent Level"]]) or 0,
            "parent": _coerce_int(r[idx["Parent ID"]]) if "Parent ID" in idx else None,
            "predecessor": _coerce_int(r[idx["Predecessor ID"]]) if "Predecessor ID" in idx else None,
            "lag": _coerce_int(r[idx["Lag"]]) or 0 if "Lag" in idx else 0,
        })

    # Compute dates via a topological walk (FS predecessor logic, 5-day workweek)
    by_id = {t["id"]: t for t in raw_tasks}
    starts: dict[int, date] = {}
    finishes: dict[int, date] = {}

    def _compute(tid: int):
        if tid in starts:
            return
        t = by_id[tid]
        # Determine start date
        if t["predecessor"] and t["predecessor"] in by_id:
            _compute(t["predecessor"])
            base = finishes[t["predecessor"]]
            s = _add_workdays(base, max(1, t["lag"] or 1))
        elif t["parent"] and t["parent"] in by_id:
            _compute(t["parent"])
            s = starts[t["parent"]]
        elif project_start:
            s = project_start
        else:
            s = date.today()
        dur = max(t["duration"], 1) if t["duration"] is not None else 1
        f = _add_workdays(s, dur - 1) if dur >= 1 else s
        starts[tid] = s
        finishes[tid] = f

    for t in raw_tasks:
        _compute(t["id"])

    # Build flat ParsedScheduleItem list, with discipline+phase context for leaf tasks
    items: list[ParsedScheduleItem] = []
    current_discipline = ""
    current_phase = ""
    total_price = 0
    for order, t in enumerate(raw_tasks):
        indent = t["indent"]
        if indent == 0:
            current_discipline = t["name"]
            current_phase = ""
        elif indent == 1:
            current_phase = t["name"]
        # tasks at indent >= 2 keep current discipline & phase
        if t["price"]:
            total_price += t["price"]
        items.append(ParsedScheduleItem(
            order_index=order,
            indent_level=min(indent, 2),
            discipline=current_discipline if indent >= 1 else "",
            phase=current_phase if indent >= 2 else "",
            task=t["name"],
            duration_days=t["duration"] or None,
            start_date=starts.get(t["id"]),
            finish_date=finishes.get(t["id"]),
            price=t["price"] or None,
            is_milestone=t["is_milestone"],
        ))

    total_duration = None
    if items:
        # Total duration = working-day span from earliest start to latest finish
        all_starts = [i.start_date for i in items if i.start_date]
        all_finishes = [i.finish_date for i in items if i.finish_date]
        if all_starts and all_finishes:
            total_duration = _workdays_between(min(all_starts), max(all_finishes))

    return ParsedSchedule(
        version=version,
        source_format="xlsx",
        source_filename=filename,
        project_name=project_name,
        project_start_date=project_start,
        total_duration_days=total_duration,
        total_price=total_price or None,
        items=items,
    )


# ============================================================
# PDF parser
# ============================================================
_DATE_RE = re.compile(r"^\d{2}/\d{2}/\d{2}$")
_DAYS_RE = re.compile(r"^\d+$")
_PRICE_RE = re.compile(r"^\$?[\d,]+$")


# Regex for a row like "30% Design 30 08/21/25 10/02/25 $30,000"
#   group 1 = name (lazy)   2 = days   3 = start   4 = finish   5 = price (optional)
_ROW_RE = re.compile(
    r"^(.+?)\s+(\d+)\s+(\d{2}/\d{2}/\d{2})\s+(\d{2}/\d{2}/\d{2})"
    r"(?:\s+\$([\d,]+))?\s*$"
)
_TOTAL_RE = re.compile(
    r"^TOTAL\s+(\d+)\s+(\d{2}/\d{2}/\d{2})\s+(\d{2}/\d{2}/\d{2})\s+\$([\d,]+)\s*$"
)
_PROJECT_INIT_PRICE_RE = re.compile(r"^Project Initiation\s+\$([\d,]+)\s*$")
_PROJECT_CLOSEOUT_PRICE_RE = re.compile(r"^Project Closeout\s+\$([\d,]+)\s*$")
_VERSION_RE = re.compile(r"-\s*(V\d+)\b")
_SAMPLE_HEADER_RE = re.compile(r"^Customer\s*:\s*(.+)$|^Company\s*:\s*(.+)$", re.IGNORECASE)


def parse_pdf_schedule(content: bytes, filename: str = "") -> ParsedSchedule:
    items: list[ParsedScheduleItem] = []
    project_name = ""
    version = "V1"
    project_start: Optional[date] = None
    total_duration: Optional[int] = None
    total_price: Optional[int] = None

    current_discipline = ""
    current_phase = ""

    with pdfplumber.open(io.BytesIO(content)) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            for raw_line in text.splitlines():
                line = raw_line.strip()
                if not line:
                    continue

                # --- Metadata ---
                vm = _VERSION_RE.search(line)
                if vm:
                    version = vm.group(1)
                # Page 1 header has "Company / Customer" pre-block; on page 2/3 it's "Customer : <name>"
                # We don't strictly need project_name from the PDF — leave blank if unsure.

                # --- Header rows we ignore ---
                low = line.lower()
                if low.startswith("project schedule") or low.startswith("task ") or low == "task":
                    continue
                if low.startswith("days start") or "days start finish" in low:
                    continue

                # --- TOTAL row ---
                m = _TOTAL_RE.match(line)
                if m:
                    total_duration = int(m.group(1))
                    project_start = project_start or _parse_mdy(m.group(2))
                    total_price = _parse_price(m.group(4))
                    continue

                # --- "Project Initiation $0" / "Project Closeout $0" milestone-only rows ---
                m = _PROJECT_INIT_PRICE_RE.match(line)
                if m:
                    current_discipline = "Project Initiation"
                    current_phase = ""
                    items.append(ParsedScheduleItem(
                        order_index=len(items), indent_level=0,
                        task="Project Initiation",
                        price=_parse_price(m.group(1)),
                        is_milestone=True,
                    ))
                    continue
                m = _PROJECT_CLOSEOUT_PRICE_RE.match(line)
                if m:
                    current_discipline = "Project Closeout"
                    current_phase = ""
                    items.append(ParsedScheduleItem(
                        order_index=len(items), indent_level=0,
                        task="Project Closeout",
                        price=_parse_price(m.group(1)),
                        is_milestone=True,
                    ))
                    continue

                # --- Standard data row ---
                m = _ROW_RE.match(line)
                if not m:
                    # Sub-items under Project Initiation appear on standalone lines with no metrics
                    # (e.g. "Contract Signed", "Deposit"). Capture them under the current discipline.
                    if current_discipline == "Project Initiation" and len(line) < 80 and not line[0].isdigit():
                        items.append(ParsedScheduleItem(
                            order_index=len(items), indent_level=1,
                            discipline=current_discipline, task=line,
                        ))
                    continue

                name = m.group(1).strip()
                days = int(m.group(2))
                start_d = _parse_mdy(m.group(3))
                finish_d = _parse_mdy(m.group(4))
                price = _parse_price(m.group(5)) if m.group(5) else None

                # Classify by name
                if name in DISCIPLINE_NAMES:
                    current_discipline = name
                    current_phase = ""
                    indent = 0
                    if not project_start and start_d:
                        project_start = start_d
                    is_milestone = True
                elif PHASE_PATTERN.match(name):
                    current_phase = name
                    indent = 1
                    is_milestone = True
                else:
                    indent = 2
                    is_milestone = False

                items.append(ParsedScheduleItem(
                    order_index=len(items), indent_level=indent,
                    discipline=current_discipline if indent >= 1 else "",
                    phase=current_phase if indent >= 2 else "",
                    task=name,
                    duration_days=days,
                    start_date=start_d, finish_date=finish_d,
                    price=price, is_milestone=is_milestone,
                ))

    return ParsedSchedule(
        version=version,
        source_format="pdf",
        source_filename=filename,
        project_name=project_name,
        project_start_date=project_start,
        total_duration_days=total_duration,
        total_price=total_price,
        items=items,
    )


# ============================================================
# Helpers
# ============================================================
def _coerce_int(v) -> Optional[int]:
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip().replace(",", "").replace("$", "")
    if not s:
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def _parse_price(s: str) -> Optional[int]:
    if not s:
        return None
    s = s.strip().replace(",", "").replace("$", "")
    if not s or not s.replace("-", "").isdigit():
        return None
    return int(s)


def _parse_mdy(s: str) -> Optional[date]:
    """MM/DD/YY → date. Two-digit year is treated as 20YY."""
    if not s or not _DATE_RE.match(s):
        return None
    m, d, y = s.split("/")
    return date(2000 + int(y), int(m), int(d))


def _coerce_date(v) -> Optional[date]:
    if v is None or v == "":
        return None
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if _DATE_RE.match(s):
        return _parse_mdy(s)
    # Try ISO
    try:
        return date.fromisoformat(s)
    except ValueError:
        return None


def _add_workdays(start: date, n_days: int) -> date:
    """Add n business days (Mon-Fri) to `start`, treating start as day 1."""
    if n_days <= 0:
        return start
    d = start
    added = 0
    while added < n_days:
        d += timedelta(days=1)
        if d.weekday() < 5:
            added += 1
    return d


def _workdays_between(a: date, b: date) -> int:
    if b < a:
        a, b = b, a
    cur = a
    count = 0
    while cur <= b:
        if cur.weekday() < 5:
            count += 1
        cur += timedelta(days=1)
    return count
