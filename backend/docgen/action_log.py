"""
Rolling action items log as Excel .xlsx. One row per action with status,
owner, due date, originating meeting, closed-in meeting, and notes.

Colors match the Castillo brand status pills.
"""
from io import BytesIO
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import BrandColors


# Status fill colors (rgb hex without #)
STATUS_FILLS = {
    "open":      ("FCE8EA", "791F1F"),   # bg, text
    "pending":   ("FDEAC0", "5E3F00"),
    "completed": ("C7E9A3", "1A3A04"),
    "cancelled": ("E6E7E8", "1A1A1A"),
}


def generate_action_items_xlsx(project, actions, output_path: Optional[Path] = None) -> bytes:
    """
    project: Project ORM object
    actions: list of ActionItem ORM objects (typically open + pending + recent closed)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Action Items"

    # ---- Title block ----
    ws["A1"] = f"Action Items — {project.name}"
    ws["A1"].font = Font(name="Helvetica", size=16, bold=True, color=BrandColors.RED.lstrip("#"))
    ws.merge_cells("A1:G1")

    if project.client:
        ws["A2"] = f"Client: {project.client.name}"
        ws["A2"].font = Font(name="Helvetica", size=10, color=BrandColors.DARK_GRAY.lstrip("#"))
        ws.merge_cells("A2:G2")

    # ---- Header row ----
    headers = ["#", "Action", "Owner", "Due", "Status", "Raised", "Closed"]
    header_row = 4
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = Font(name="Helvetica", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=BrandColors.RED.lstrip("#"))
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # ---- Data ----
    thin_border = Border(
        left=Side(style="thin", color="BCBEC0"),
        right=Side(style="thin", color="BCBEC0"),
        top=Side(style="thin", color="BCBEC0"),
        bottom=Side(style="thin", color="BCBEC0"),
    )

    for idx, action in enumerate(actions, start=1):
        row = header_row + idx
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=action.text)
        ws.cell(row=row, column=3, value=action.owner or "")
        ws.cell(row=row, column=4, value=action.due_date.strftime("%m/%d/%Y") if action.due_date else "")
        status_cell = ws.cell(row=row, column=5, value=action.status.capitalize())
        bg, fg = STATUS_FILLS.get(action.status, ("FFFFFF", "1A1A1A"))
        status_cell.fill = PatternFill("solid", fgColor=bg)
        status_cell.font = Font(name="Helvetica", size=10, bold=True, color=fg)
        ws.cell(row=row, column=6, value=action.originating_meeting.meeting_date.strftime("%m/%d/%Y") if action.originating_meeting else "")
        ws.cell(row=row, column=7, value=action.closed_in_meeting.meeting_date.strftime("%m/%d/%Y") if action.closed_in_meeting else "")

        for col in range(1, len(headers) + 1):
            c = ws.cell(row=row, column=col)
            c.border = thin_border
            if not c.font.bold:
                c.font = Font(name="Helvetica", size=10)
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # ---- Column widths ----
    widths = [5, 50, 12, 12, 14, 12, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A5"

    # ---- Save ----
    buf = BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    if output_path:
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        Path(output_path).write_bytes(data)
    return data
