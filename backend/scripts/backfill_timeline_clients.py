"""Tie existing timeline PROJECTS to their client by matching the project name
to the Contracts sheet's Project Site -> Company Name (2024+ rows only). Only
fills projects whose client is currently blank, and only on confident matches:
  - single name: exact site match, or the project name is the leading word(s)
    of a site (e.g. "Trigo" -> site "Trigo Sol").
  - "&" / "/" bundles: every part must match AND agree on one company
    (e.g. "Nesler & E1300" -> both Utopian Power).
Ambiguous / unmatched projects are left blank and reported.

Dry-run by default; pass --apply to write.
    python scripts/backfill_timeline_clients.py ["...Contracts Sheet.xlsx"] [--apply]
"""
from __future__ import annotations
import os, re, sys, datetime
BACKEND = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if BACKEND not in sys.path:
    sys.path.insert(0, BACKEND)
import config  # noqa: F401
import openpyxl
from db.session import get_session_factory
from db.models import TimelineProject

DEFAULT_XLSX = os.path.join(os.path.expanduser("~"), "Downloads", "Contracts Sheet.xlsx")


def _year(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.year
    m = re.search(r"(20\d{2})", str(v) if v is not None else "")
    return int(m.group(1)) if m else None


def _norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", (s or "").lower()).strip()


def build_site_map(path: str) -> dict[str, str]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["Contracts Sheet"] if "Contracts Sheet" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    H = [str(c).strip() if c is not None else "" for c in next(rows)]
    ci, si, di = H.index("Company Name"), H.index("Project Site"), H.index("Project Creation Date")
    site2co: dict[str, str] = {}
    for r in rows:
        co = r[ci] if ci < len(r) else None
        yr = _year(r[di] if di < len(r) else None)
        if not co or yr is None or yr < 2024:
            continue
        ns = _norm(r[si] if si < len(r) else None)
        if ns:
            site2co.setdefault(ns, str(co).strip())
    return site2co


def matcher(site2co: dict[str, str]):
    def one(part: str):
        if len(part) < 3:
            return None
        if part in site2co:
            return site2co[part]
        for ns, co in site2co.items():
            if ns.startswith(part + " "):
                return co
        return None

    def match(name: str):
        parts = [_norm(x) for x in re.split(r"[&/]", name) if _norm(x)]
        if len(parts) > 1:
            cos = [one(p) for p in parts]
            if all(cos) and len(set(cos)) == 1:
                return cos[0]
            return None
        return one(_norm(name))

    return match


def main() -> None:
    args = [a for a in sys.argv[1:] if a != "--apply"]
    apply = "--apply" in sys.argv
    path = args[0] if args else DEFAULT_XLSX
    match = matcher(build_site_map(path))
    db = get_session_factory()()
    try:
        set_n = skip_n = 0
        for p in sorted(db.query(TimelineProject).all(), key=lambda x: x.name.lower()):
            if p.client:  # never clobber an existing client
                continue
            co = match(p.name)
            if co:
                set_n += 1
                print(f"  SET  {p.name!r:42} -> {co}")
                if apply:
                    p.client = co
            else:
                skip_n += 1
                print(f"  skip {p.name!r:42}")
        if apply:
            db.commit()
        print(f"\n{'APPLIED' if apply else 'DRY-RUN'}: {set_n} projects matched, {skip_n} left blank.")
    finally:
        db.close()


if __name__ == "__main__":
    main()
