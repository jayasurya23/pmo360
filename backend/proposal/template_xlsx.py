"""Desktop "Save Template" / "Load Template" .xlsx round-trip.

A faithful 1:1 port of Full_proposal_V9.py ``save_template_excel`` (7505-7677)
and ``load_template_excel`` (7680-7915) onto the framework-free proposal package
(no FastAPI / ORM import — only openpyxl + the sibling proposal modules).

    write_template_xlsx(tree_json, info_json, config_json) -> bytes
        three sheets — Project Info, Tasks (18 cols), Split Deposit Memory —
        in the desktop's exact column order, so a file we write loads in the
        Tkinter tool and vice-versa.

    read_template_xlsx(content: bytes) -> dict
        {tree_json, info_json, config_json, last_task_id}

The read path does NOT call ``build_tree``: the tree is reconstructed directly
from the Tasks rows by ID / Parent ID (the same two-pass parent-wiring loop as
``deserialize_tree``), so the already-baked Client-Review / Record-Drawings
nodes are not duplicated. Dates are recomputed by the caller via
``calculate_all_dates`` (the Tasks sheet stores no start/end columns — desktop
parity).
"""
from __future__ import annotations

import io
import math
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .models import ProposalItem, ProjectInfo
from .serialization import (
    serialize_tree, deserialize_tree, build_info_json, _to_mdy,
)

# The desktop's exact 18-column Tasks order (Full_proposal_V9.py:7556-7560).
# Predecessor ID + Lag sit between "Price Only" and "Is Start Pinned" — keep
# this order so a desktop-saved file round-trips through here unchanged.
TASK_HEADERS = (
    "ID", "Name", "Duration", "Targeted Hours", "Price", "Is Milestone",
    "Indent Level", "Enabled", "Price Only", "Predecessor ID", "Lag",
    "Is Start Pinned", "Parent ID", "Type", "Type User Set",
    "Show Start Date", "Show End Date", "Task Utilization",
)


# ---------------------------------------------------------------------------
# Small private helpers
# ---------------------------------------------------------------------------
def _opt_float(v):
    """None / "" -> None, else float (best effort, None on failure)."""
    if v is None or v == "":
        return None
    try:
        return float(str(v))
    except (TypeError, ValueError):
        return None


def _valid_iso(s: str) -> bool:
    """True when ``s`` parses as a YYYY-MM-DD date."""
    try:
        datetime.strptime(s, "%Y-%m-%d")
        return True
    except (TypeError, ValueError):
        return False


def _assign_missing_ids(items) -> None:
    """Belt-and-suspenders: guarantee every node has an int id (+ parent_id).

    The export reads the server-persisted tree where every node already has an
    int id, but a hand-edited / freshly-added node could carry ``id is None``.
    The desktop loader fails on a null ID join key, so we assign sequential ids
    (pre-order) starting at max(existing)+1 before writing, and fix children's
    parent_id to match.
    """
    existing = []

    def collect(nodes):
        for it in nodes:
            if it.id is not None:
                existing.append(int(it.id))
            collect(it.children)

    collect(items)
    nxt = (max(existing) + 1) if existing else 1

    def fix(nodes, parent_id):
        nonlocal nxt
        for it in nodes:
            if it.id is None:
                it.id = nxt
                nxt += 1
            it.parent_id = parent_id
            fix(it.children, it.id)

    fix(items, None)


# ---------------------------------------------------------------------------
# Export — tree/info/config JSON -> .xlsx bytes
# ---------------------------------------------------------------------------
def write_template_xlsx(*, tree_json, info_json, config_json) -> bytes:
    """Serialize a persisted proposal version into a desktop-shaped template."""
    tree_json = tree_json or []
    info_json = info_json or {}
    config_json = config_json or {}

    items, _id_map = deserialize_tree(tree_json)
    _assign_missing_ids(items)

    # Flatten depth-first pre-order (parent then its children), tracking the
    # parent's id — verbatim port of _flatten_items (7569-7596).
    flat: list[tuple[ProposalItem, "int | None"]] = []

    def _flatten(nodes, parent_id=None):
        for it in nodes:
            flat.append((it, parent_id))
            if it.children:
                _flatten(it.children, parent_id=it.id)

    _flatten(items)

    sdm = info_json.get("split_deposit_memory") or {}

    wb = Workbook()

    # --- Sheet 1: Project Info -------------------------------------------
    info_ws = wb.active
    info_ws.title = "Project Info"

    # "" (our blank == bundled Castillo logo) is persisted as the desktop
    # sentinel so the file stays openable in the Tkinter tool.
    logo_path = info_json.get("logo_path", "") or ""
    logo_to_save = "DEFAULT_LOGO" if not logo_path else logo_path

    last_task_id = max((int(it.id) for it, _ in flat if it.id is not None),
                       default=0)

    project_data = {
        "Project Name": info_json.get("project_title", ""),
        "Company Name": info_json.get("customer_name", ""),
        "Proposal Version": (info_json.get("version") or "V1").strip(),
        "Project Start Date": config_json.get("project_start", ""),
        "Logo Path": logo_to_save,
        "Client Logo Path": info_json.get("client_logo_path", ""),
        "Last Task ID": last_task_id,
        "Project Location": info_json.get("project_location", ""),
        "Project State": info_json.get("project_state", ""),
        "Project Size (MW)": info_json.get("project_size_mw", ""),
        "Deposit Percentage": sdm.get(
            "deposit_percentage", info_json.get("deposit_percent", 30.0)),
        "Custom Holidays": ",".join(sorted(config_json.get("custom_holidays") or [])),
        "Utilization Percent": config_json.get("utilization_percent", 100.0),
    }

    info_ws.append(["Attribute", "Value"])
    for key, value in project_data.items():
        info_ws.append([key, value])
    info_ws["A1"].font = Font(bold=True)
    info_ws["B1"].font = Font(bold=True)
    info_ws.column_dimensions["A"].width = 20
    info_ws.column_dimensions["B"].width = 50

    # --- Sheet 2: Tasks ---------------------------------------------------
    tasks_ws = wb.create_sheet(title="Tasks")
    tasks_ws.append(list(TASK_HEADERS))
    for cell in tasks_ws[1]:
        cell.font = Font(bold=True)

    for it, parent_id in flat:
        tasks_ws.append([
            it.id,
            it.name,
            it.duration,
            it.targeted_hours,
            it.price,
            bool(it.is_milestone),
            it.indent_level,
            bool(it.enabled),
            bool(it.price_only),
            it.predecessor_id,
            it.lag,
            bool(it.is_start_pinned),
            parent_id,
            it.predecessor_type,           # FS / SS / FF / SF
            bool(it.predecessor_type_user_set),
            bool(it.show_start_date),
            bool(it.show_end_date),
            it.task_utilization,
        ])

    # Auto-size each column (header width floor 10) — desktop parity 7606-7613.
    for col_idx, header in enumerate(TASK_HEADERS, 1):
        letter = get_column_letter(col_idx)
        max_width = max(len(str(header)), 10)
        for row in tasks_ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_width = max(max_width, len(str(cell.value)))
        tasks_ws.column_dimensions[letter].width = max_width + 2

    # --- Sheet 3: Split Deposit Memory -----------------------------------
    split_ws = wb.create_sheet(title="Split Deposit Memory")
    split_ws.append(["Key", "Value"])
    for cell in split_ws[1]:
        cell.font = Font(bold=True)

    # Group A — scalars (always written, even when split was never applied).
    split_ws.append(["deposit_percentage", sdm.get("deposit_percentage", 0.0)])
    split_ws.append(["has_been_applied", bool(sdm.get("has_been_applied", False))])
    split_ws.append([
        "selected_task_indices",
        ",".join(str(x) for x in sdm.get("selected_task_keys", [])),
    ])
    split_ws.append(["split_mode", sdm.get("split_mode", "percent_of_each")])
    split_ws.append(["deposit_target", sdm.get("deposit_target", "deposit")])

    # Group B — per-task split percentages.
    for tid, pct in (sdm.get("task_percentages") or {}).items():
        split_ws.append([f"task_percentage_{tid}", pct])

    # Group C — pre-split original task prices.
    for tid, price in (sdm.get("original_task_prices") or {}).items():
        split_ws.append([f"original_price_{tid}", price])

    # Group D — pre-split prices of the target items (Deposit / DD).
    for iid, price in (sdm.get("original_target_prices") or {}).items():
        split_ws.append([f"original_target_price_{iid}", price])

    # Group E — "template original price" for every non-milestone item, for the
    # desktop's full-reset feature. The web split model has no separate template
    # store, so synthesize it from original_task_prices (else current price) for
    # file-format / desktop interop. These rows are ignored on import.
    orig = sdm.get("original_task_prices") or {}
    for it, _ in flat:
        if not it.is_milestone and it.id is not None:
            split_ws.append([
                f"template_original_price_{it.id}",
                orig.get(str(it.id), it.price),
            ])

    split_ws.column_dimensions["A"].width = 25
    split_ws.column_dimensions["B"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Import — .xlsx bytes -> tree/info/config JSON
# ---------------------------------------------------------------------------
def read_template_xlsx(content: bytes) -> dict:
    """Parse a saved template workbook back into JSON blobs.

    Returns ``{tree_json, info_json, config_json, last_task_id}``. Raises
    ``ValueError`` when the required sheets are missing (caller maps to 400).
    """
    wb = load_workbook(io.BytesIO(content), data_only=True)
    if "Project Info" not in wb.sheetnames or "Tasks" not in wb.sheetnames:
        raise ValueError("Workbook is missing the 'Project Info' / 'Tasks' sheet")

    # --- (1) Project Info -> info_json + config inputs --------------------
    info_ws = wb["Project Info"]
    pi = {row[0]: row[1] for row in info_ws.iter_rows(min_row=2, values_only=True)
          if row and row[0] is not None}

    project_start = _to_mdy(pi.get("Project Start Date")) or ""
    util = float(pi.get("Utilization Percent", 100.0) or 100.0)
    custom_hols = [t.strip() for t in str(pi.get("Custom Holidays") or "").split(",")
                   if _valid_iso(t.strip())]
    last_task_id = int(pi.get("Last Task ID", 0) or 0)

    logo_raw = pi.get("Logo Path")
    logo_path = "" if logo_raw in (None, "", "DEFAULT_LOGO") else str(logo_raw)

    pinfo = ProjectInfo(
        project_title=str(pi.get("Project Name") or "").strip(),
        customer_name=str(pi.get("Company Name") or "").strip(),
        project_location=str(pi.get("Project Location") or "").strip(),
        project_state=str(pi.get("Project State") or "").strip(),
        project_size_mw=("" if pi.get("Project Size (MW)") in (None, "")
                         else str(pi.get("Project Size (MW)"))),
        version=str(pi.get("Proposal Version") or "V1").strip() or "V1",
        logo_path=logo_path,
        client_logo_path=str(pi.get("Client Logo Path") or ""),
    )
    info = {
        "deposit_percent": float(pi.get("Deposit Percentage", 30.0) or 30.0),
        "date": project_start,
        "logo_path": logo_path,
        "client_logo_path": pinfo.client_logo_path,
    }
    info_json = build_info_json(info, pinfo)

    # --- (2) Tasks -> tree (two-pass, NO node injection) ------------------
    tasks_ws = wb["Tasks"]
    headers = [c.value for c in tasks_ws[1]]

    rows: list[dict] = []
    items_by_id: dict[int, ProposalItem] = {}
    for raw in tasks_ws.iter_rows(min_row=2, values_only=True):
        if raw is None or raw[0] is None:          # skip blank/trailing rows
            continue
        d = dict(zip(headers, raw))
        rows.append(d)
        iid = int(d["ID"])
        name = d.get("Name") or ""
        type_user_set = bool(d.get("Type User Set", False))
        # Mirror the loader's default_type then explicit-Type override (7770-7771)
        default_type = "FF" if "study" in name.lower() else "FS"
        pred_type = d.get("Type") or default_type
        pred_raw = d.get("Predecessor ID")
        parent_raw = d.get("Parent ID")
        enabled_raw = d.get("Enabled")
        sstart_raw = d.get("Show Start Date")
        send_raw = d.get("Show End Date")

        it = ProposalItem(
            name=name,
            duration=math.ceil(float(d.get("Duration", 0) or 0)),
            price=int(float(d.get("Price", 0) or 0)),
            is_milestone=bool(d.get("Is Milestone")),
            indent_level=int(d.get("Indent Level", 0) or 0),
            id=iid,
            predecessor_id=(int(pred_raw) if pred_raw is not None else None),
            # set user-set BEFORE pred_type matters: __post_init__ only overrides
            # predecessor_type from the name when user_set is False.
            predecessor_type=pred_type,
            predecessor_type_user_set=type_user_set,
            lag=int(d.get("Lag", 0) or 0),
            targeted_hours=d.get("Targeted Hours"),
            is_start_pinned=bool(d.get("Is Start Pinned", False)),
            enabled=(bool(enabled_raw) if enabled_raw is not None else True),
            price_only=bool(d.get("Price Only", False)),
            show_start_date=(bool(sstart_raw) if sstart_raw is not None else True),
            show_end_date=(bool(send_raw) if send_raw is not None else True),
            task_utilization=_opt_float(d.get("Task Utilization")),
            parent_id=(int(parent_raw) if parent_raw is not None else None),
        )
        items_by_id[iid] = it

    # Pass 2 — wire parents (verbatim deserialize_tree loop; dangling parent ->
    # root, matching the desktop's else branch 7805-7806).
    roots: list[ProposalItem] = []
    for d in rows:
        iid = int(d["ID"])
        cur = items_by_id[iid]
        parent_raw = d.get("Parent ID")
        pid = int(parent_raw) if parent_raw is not None else None
        cur.parent_id = pid
        if pid is not None and pid in items_by_id:
            parent = items_by_id[pid]
            parent.children.append(cur)
            cur.parent = parent
        else:
            roots.append(cur)

    # --- (3) Split Deposit Memory -> info_json["split_deposit_memory"] ----
    sdm = {
        "deposit_percentage": 30.0,
        "split_mode": "percent_of_each",
        "deposit_target": "deposit",
        "has_been_applied": False,
        "task_percentages": {},
        "original_task_prices": {},
        "original_target_prices": {},
        "selected_task_keys": [],
    }
    try:
        if "Split Deposit Memory" in wb.sheetnames:
            split_ws = wb["Split Deposit Memory"]
            for row in split_ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 2 or row[0] is None:
                    continue
                key, value = row[0], row[1]
                if key == "deposit_percentage":
                    sdm["deposit_percentage"] = float(value) if value is not None else 0.0
                elif key == "has_been_applied":
                    sdm["has_been_applied"] = bool(value) if value is not None else False
                elif key == "selected_task_indices":
                    if value is not None and str(value).strip():
                        # web selection keys are String(id); keep as strings
                        sdm["selected_task_keys"] = [
                            x.strip() for x in str(value).split(",") if x.strip()
                        ]
                elif key == "split_mode":
                    if value:
                        sdm["split_mode"] = str(value)
                elif key == "deposit_target":
                    tv = str(value).strip().lower() if value else "deposit"
                    sdm["deposit_target"] = tv if tv in ("deposit", "due_diligence") else "deposit"
                elif isinstance(key, str) and key.startswith("task_percentage_"):
                    tid = key[len("task_percentage_"):]
                    sdm["task_percentages"][str(int(tid))] = float(value) if value is not None else 0.0
                elif isinstance(key, str) and key.startswith("original_target_price_"):
                    iid = key[len("original_target_price_"):]
                    sdm["original_target_prices"][str(int(iid))] = float(value) if value is not None else 0.0
                elif isinstance(key, str) and key.startswith("original_price_"):
                    tid = key[len("original_price_"):]
                    sdm["original_task_prices"][str(int(tid))] = float(value) if value is not None else 0.0
                # template_original_price_* intentionally ignored (no web home)
    except Exception:  # noqa: BLE001 — bad split sheet -> keep defaults (desktop parity)
        sdm = {
            "deposit_percentage": 30.0,
            "split_mode": "percent_of_each",
            "deposit_target": "deposit",
            "has_been_applied": False,
            "task_percentages": {},
            "original_task_prices": {},
            "original_target_prices": {},
            "selected_task_keys": [],
        }

    # Prune split data to ids that still exist (port of _validate_split_deposit_data;
    # web keys are ids, so filter by id rather than the desktop's positional index).
    current = {str(i) for i in items_by_id}
    sdm["task_percentages"] = {k: v for k, v in sdm["task_percentages"].items() if k in current}
    sdm["original_task_prices"] = {k: v for k, v in sdm["original_task_prices"].items() if k in current}
    sdm["original_target_prices"] = {k: v for k, v in sdm["original_target_prices"].items() if k in current}
    sdm["selected_task_keys"] = [k for k in sdm["selected_task_keys"] if k in current]
    info_json["split_deposit_memory"] = sdm

    return {
        "tree_json": serialize_tree(roots),
        "info_json": info_json,
        "config_json": {
            "project_start": project_start,
            "utilization_percent": util,
            "fs_start_next_day": False,
            "disabled_holidays": [],
            "custom_holidays": custom_hols,
        },
        "last_task_id": last_task_id,
    }
