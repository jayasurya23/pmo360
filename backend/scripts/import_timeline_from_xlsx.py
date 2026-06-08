"""Import the live capacity plan from the 'Refresh' sheet of the Timeline
estimator workbook into the Timeline Estimator tables.

Each engineer row -> a TimelineResource (discipline from title/section,
placeholder flagged). Each merged/standalone weekly cell -> a
TimelineAssignment, with labels like "Highland 90" / "Highland IFC" grouped
under one TimelineProject ("Highland") and a parsed milestone.

By default this WIPES the three timeline tables and reloads from the sheet
(the workbook is the source of truth). Pass --keep to append instead.

Usage:
    python scripts/import_timeline_from_xlsx.py "<path.xlsx>" [--keep] [--sheet Refresh]
"""
from __future__ import annotations

import os
import re
import sys
from datetime import datetime, date, timedelta

import openpyxl

BACKEND = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if BACKEND not in sys.path:
    sys.path.insert(0, BACKEND)

import config  # noqa: F401  -- loads .env
from db.session import get_engine, get_session_factory
from db.models import Base, TimelineResource, TimelineProject, TimelineAssignment

# ---- discipline mapping ------------------------------------------------
# Explicit title keywords win; otherwise fall back to the row-section the
# engineer sits in (the sheet groups Electrical -> Structural -> Civil ->
# Water down column B).
def disc_from_title(title: str | None) -> str | None:
    t = (title or "").lower()
    if not t:
        return None
    if t.startswith("ee") or "electric" in t:
        return "Electrical"
    if t.startswith("ce") or "civil" in t:
        return "Civil"
    if "struct" in t:
        return "Structural"
    if "water" in t:
        return "Water"
    return None  # Intern / Associate / blank -> use section


# Row sections in the Refresh sheet (validated against the parse).
SECTIONS = [
    (4, 22, "Electrical"),
    (23, 25, "Structural"),
    (26, 30, "Civil"),
    (31, 99, "Water"),
]
def disc_from_section(row: int) -> str:
    for lo, hi, d in SECTIONS:
        if lo <= row <= hi:
            return d
    return "Electrical"


PLACEHOLDER_HINTS = ("new hire", "solvida", "gers", "finulent", "structurology", "tbd", "vendor")
def is_placeholder(name: str) -> bool:
    n = name.lower()
    return any(h in n for h in PLACEHOLDER_HINTS)


# ---- label -> (project_name, milestone, status) ------------------------
PHASE_RE = re.compile(
    r"^(?P<name>.+?)[\s\-]*\b(?P<phase>IFC|IFP|Stage\s*B|Studies|90%?|60%?|30%?|10%?)\s*$",
    re.IGNORECASE,
)
PHASE_NORM = {"90": "90%", "60": "60%", "30": "30%", "10": "10%", "stage b": "Stage B",
              "ifc": "IFC", "ifp": "IFP", "studies": "Studies"}
STATUS_BY_LABEL = {"ooo": "on_hold", "tentative": "not_contracted", "hold": "on_hold"}


def clean(s: str) -> str:
    return re.sub(r"\s+", " ", str(s).replace("\n", " ")).strip()


def parse_label(raw: str):
    label = clean(raw)
    status = STATUS_BY_LABEL.get(label.lower())
    # only split a milestone off SIMPLE single-project labels (no & / , / ()).
    if not re.search(r"[&,/()]", label):
        m = PHASE_RE.match(label)
        if m:
            name = clean(m.group("name"))
            phase = m.group("phase").lower().replace(" ", "")
            phase = PHASE_NORM.get(phase, m.group("phase"))
            if name:
                return name, phase, status
    return label, None, status


def find_header_row(ws):
    best_row, best = None, []
    for r in range(1, 12):
        dates = [(c, ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)
                 if isinstance(ws.cell(row=r, column=c).value, (datetime, date))]
        if len(dates) > len(best):
            best_row, best = r, dates
    return best_row, best


def build_merged_map(ws):
    merged = {}
    for rng in ws.merged_cells.ranges:
        tl = ws.cell(row=rng.min_row, column=rng.min_col)
        for rr in range(rng.min_row, rng.max_row + 1):
            for cc in range(rng.min_col, rng.max_col + 1):
                merged[(rr, cc)] = (rng.min_col, rng.max_col, rng.min_row, tl.value)
    return merged


def main():
    args = sys.argv[1:]
    keep = "--keep" in args
    args = [a for a in args if a != "--keep"]
    sheet = "Refresh"
    if "--sheet" in args:
        i = args.index("--sheet"); sheet = args[i + 1]; del args[i:i + 2]
    path = args[0] if args else r"C:\Users\jbhaskar\Downloads\Timeline estimator.xlsx"

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    header_row, header_dates = find_header_row(ws)
    week_cols = sorted(c for c, _ in header_dates)
    first_col, last_col = week_cols[0], week_cols[-1]
    col_date = {c: (v.date() if hasattr(v, "date") else v) for c, v in header_dates}
    week_len = timedelta(days=4)  # Mon..Fri inclusive end
    merged = build_merged_map(ws)

    Base.metadata.create_all(get_engine())
    Session = get_session_factory()
    db = Session()
    try:
        if not keep:
            db.query(TimelineAssignment).delete()
            db.query(TimelineProject).delete()
            db.query(TimelineResource).delete()
            db.flush()

        proj_by_name: dict[str, TimelineProject] = {}
        used_names: dict[str, int] = {}
        n_res = n_assign = 0

        for r in range(header_row + 1, ws.max_row + 1):
            name = ws.cell(row=r, column=2).value
            if name in (None, ""):
                continue
            name = clean(name)
            title = ws.cell(row=r, column=3).value
            title = clean(title) if title else None
            disc = disc_from_title(title) or disc_from_section(r)

            # de-dup resource display names (e.g. multiple "New Hire")
            disp = name
            if name in used_names:
                used_names[name] += 1
                disp = f"{name} {used_names[name]}"
            else:
                used_names[name] = 1

            res = TimelineResource(
                name=disp, discipline=disc, title=title,
                is_placeholder=is_placeholder(name), active=True,
                order_index=r,
            )
            db.add(res); db.flush(); n_res += 1

            # walk the week columns -> bars
            c = first_col
            while c <= last_col:
                m = merged.get((r, c))
                if m and m[2] == r and m[1] >= first_col:
                    lo, hi, _, val = m
                    lo = max(lo, first_col); hi = min(hi, last_col)
                    step_to = hi + 1
                else:
                    lo = hi = c; val = ws.cell(row=r, column=c).value; step_to = c + 1
                c = step_to
                if val in (None, "") or isinstance(val, (datetime, date)):
                    continue
                project_name, milestone, status = parse_label(val)
                if not project_name:
                    continue
                key = project_name.lower()
                proj = proj_by_name.get(key)
                if proj is None:
                    proj = TimelineProject(name=project_name, status="in_progress")
                    db.add(proj); db.flush()
                    proj_by_name[key] = proj
                db.add(TimelineAssignment(
                    timeline_project_id=proj.id, resource_id=res.id,
                    discipline=disc, milestone=milestone,
                    start_date=col_date[lo], end_date=col_date[hi] + week_len,
                    utilization=1.0, status=status, label=clean(val),
                    order_index=0,
                ))
                n_assign += 1

        db.commit()
        print(f"Imported: {n_res} resources, {len(proj_by_name)} projects, {n_assign} assignments "
              f"from sheet '{sheet}'. Weeks {col_date[first_col]} .. {col_date[last_col]}.")
    finally:
        db.close()


if __name__ == "__main__":
    main()
