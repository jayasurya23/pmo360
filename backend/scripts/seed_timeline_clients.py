"""Resync timeline_clients from the Contracts sheet — only companies that have
a Project Creation Date in 2024 or later. Idempotent: removes previously-seeded
companies that no longer qualify, keeps user-added clients untouched.

Usage (path optional; defaults to the Downloads copy):
    python scripts/seed_timeline_clients.py ["C:\\path\\Contracts Sheet.xlsx"]
"""
from __future__ import annotations
import os, re, sys, datetime
BACKEND = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if BACKEND not in sys.path:
    sys.path.insert(0, BACKEND)
import config  # noqa: F401
import openpyxl
from db.session import get_engine, get_session_factory
from db.models import Base, TimelineClient

DEFAULT_XLSX = os.path.join(os.path.expanduser("~"), "Downloads", "Contracts Sheet.xlsx")
SKIP = {"none", "n/a", "na", "tbd", "-", ""}
MIN_YEAR = 2024


def _year(v) -> int | None:
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.year
    m = re.search(r"(20\d{2})", str(v) if v is not None else "")
    return int(m.group(1)) if m else None


def companies_2024_plus(path: str) -> list[str]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["Contracts Sheet"] if "Contracts Sheet" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(rows)]
    ci = header.index("Company Name")
    di = header.index("Project Creation Date")
    seen: dict[str, str] = {}
    for r in rows:
        name = r[ci] if ci < len(r) else None
        yr = _year(r[di] if di < len(r) else None)
        if name is None or yr is None or yr < MIN_YEAR:
            continue
        s = str(name).strip()
        if s and s.lower() not in SKIP:
            seen.setdefault(s.lower(), s)
    return sorted(seen.values(), key=str.lower)


def main() -> None:
    path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    keep = companies_2024_plus(path)
    Base.metadata.create_all(get_engine())
    db = get_session_factory()()
    try:
        # Drop previously seed-sourced rows (created_by_id IS NULL); keep user adds.
        removed = db.query(TimelineClient).filter(TimelineClient.created_by_id.is_(None)).delete()
        db.flush()
        have = {c.name.strip().lower() for c in db.query(TimelineClient).all()}
        added = 0
        for name in keep:
            if name.lower() not in have:
                db.add(TimelineClient(name=name))
                have.add(name.lower())
                added += 1
        db.commit()
        print(f"2024+ companies: {len(keep)} | removed {removed} old seeded | added {added} | total {len(have)}.")
    finally:
        db.close()


if __name__ == "__main__":
    main()
